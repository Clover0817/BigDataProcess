import openpyxl
import math

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']
count = 0
totalList = []

row_id = 1
for row in ws:
        if row_id != 1:
                sum_v = ws.cell(row = row_id, column = 3).value * 0.3
                sum_v += ws.cell(row = row_id, column = 4).value * 0.35
                sum_v += ws.cell(row = row_id, column = 5).value * 0.34
                sum_v += ws.cell(row = row_id, column = 6).value
                sum_v = round(sum_v, 2)
                ws.cell(row = row_id, column = 7).value = sum_v
                totalList.append(sum_v)
                count += 1
        row_id += 1

if count > 0:
	gradeDic = {}
	totalList.sort(reverse=True)
	for i in range(count):
       		gradeDic[totalList[i]] = i + 1

	gradeList = list(gradeDic.values())
	for i in totalList:
        	if gradeDic[i] <= math.trunc(count * 0.3):
               		if gradeDic[i] <= math.trunc(count * 0.3 * 0.5):
                        	gradeDic[i] = 'A+'
                	else:
                        	gradeDic[i] = 'A0'
        	elif gradeDic[i] <= math.trunc(count * 0.7):
                	if gradeDic[i] <= math.trunc((count * 0.7 + count * 0.3) * 0.5):
                        	gradeDic[i] = 'B+'
                	else:
                        	gradeDic[i] = 'B0'
        	else:
                	if gradeDic[i] <= math.trunc((count * 1.0 + count * 0.7) * 0.5):
                        	gradeDic[i] = 'C+'
                	else:
                        	gradeDic[i] = 'C0'

row_id = 1
for row in ws:
	if row_id != 1:
		for total in totalList:
			if ws.cell(row = row_id, column = 7).value == total:
				ws.cell(row = row_id, column = 8).value = gradeDic[total]
	row_id += 1
wb.save("student.xlsx")

